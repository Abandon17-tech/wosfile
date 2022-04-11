import glob
import os

import wosfile
import xlwt
import openpyxl as xl
from collections import Counter


class Paper:
    def __init__(self, wos):
        self.wos = wos  # UT
        self.doi = "n/a"  # DI
        self.pm = "n/a"  # PM
        self.title = ""  # TI
        self.link = ""
        self.authors = []  # AF
        self.source = ""  # SO
        self.research_area = ""  # SC
        self.type = ""  # DT
        self.volume = "n/a"  # VL
        self.issue = "n/a"  # IS
        self.page = "n/a"  # PG
        self.year = "n/a"  # PY


def read_paper(file_name):
    print(file_name)
    files = glob.glob(file_name)
    paper_list = []

    i = 0
    for rec in wosfile.records_from(files):
        # print(i, " ", rec['UT'])
        p = Paper(rec['UT'])
        if 'DI' in rec.keys():
            p.doi = rec['DI']
        if 'PM' in rec.keys():
            p.pm = rec['PM']
        p.title = rec['TI']
        p.authors = rec['AF']
        p.source = rec['SO']
        p.research_area = rec['SC']
        p.type = rec['DT']
        if 'VL' in rec.keys():
            p.volume = rec['VL']
        if 'IS' in rec.keys():
            p.issue = rec['IS']
        if 'PG' in rec.keys():
            p.page = rec['PG']
        if 'PY' in rec.keys():
            p.year = rec['PY']

        paper_list.append(p)
        i += 1

    return paper_list


def save_file_paper(file_name, paper_list):
    title = ["入藏号", "DOI", "Pubmed ID", "论文标题", "链接", "作者"]

    f_statistic = xlwt.Workbook()  # 创建工作簿
    statSheet = f_statistic.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet

    for i in range(len(title)):
        statSheet.write(0, i, title[i])

    for i in range(len(paper_list)):
        paper = paper_list[i]

        statSheet.write(i + 1, 0, paper.wos)
        statSheet.write(i + 1, 1, paper.doi)
        statSheet.write(i + 1, 2, paper.title)
        statSheet.write(i + 1, 3, paper.doi)
        statSheet.write(i + 1, 4, paper.link)
        statSheet.write(i + 1, 5, paper.authors)

    f_statistic.save(file_name)


def write_excel_file(result_path, paper_list):
    print('***** 开始写入excel文件 ' + result_path + ' ***** \n')
    if os.path.exists(result_path):
        print('***** excel已存在，在表后添加数据 ' + result_path + ' ***** \n')
        workbook = xl.load_workbook(result_path)
    else:
        print('***** excel不存在，创建excel ' + result_path + ' ***** \n')
        workbook = xl.Workbook()
        workbook.save(result_path)

    sheet = workbook.active

    for i in range(len(paper_list)):
        paper = paper_list[i]
        data = [str(paper.wos), str(paper.doi), str(paper.title), str(paper.doi), str(paper.link),
                '; '.join(paper.authors), str(paper.source), str(paper.research_area), str(paper.type),
                str(paper.volume), str(paper.issue), str(paper.page), str(paper.year)]
        sheet.append(data)

    workbook.save(result_path)
    print('***** 生成Excel文件 ' + result_path + ' ***** \n')
